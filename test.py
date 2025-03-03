import cv2
from matplotlib import pyplot as plt
from torch.utils.data import DataLoader
import openpyxl
from net import *
from data import *
import torch
import numpy as np
import tqdm
import os
os.environ['KMP_DUPLICATE_LIB_OK']='True'


def calculate_sem_seg_metrics(pred, target, num_classes):
    """
    计算语义分割评估指标的实现函数
    """
    pred = pred.data.cpu().numpy()
    target = target.data.cpu().numpy()

    num_colors, height, width = pred.shape
    pred_ = np.zeros([height*width],dtype=np.int64)
    target_ = np.zeros([height*width],dtype=np.int64)
    t=0
    for y in range(height):
        for x in range(width):
            pred_pix = pred[:, y, x]
            target_pix = target[:, y, x]
            pred_[t] = np.argmax(pred_pix)
            target_[t] = np.argmax(target_pix)
            t=t+1
    """
        计算语义分割任务的评估指标

        参数:
        pred (numpy.ndarray): 模型预测结果,形状为 (num_classes, H, W)
        target (numpy.ndarray): ground truth标签,形状为 (num_classes, H, W)
        num_classes (int): 分割类别数

        返回:
        precision (float): 精确率
        pa (float): 像素准确率
        miou (float): 平均IoU
        fwiou (float): 频权IoU
        f1 (float): F1分数
    """
    # 展平预测结果和标签数据
    pred = pred_.reshape(-1)
    target = target_.reshape(-1)

    # 初始化评估指标
    total_pixels = pred.size
    positive_pixels = np.zeros(num_classes)
    true_positive_pixels = np.zeros(num_classes)
    negative_pixels = np.zeros(num_classes)

    # 计算各类别的像素数量
    for cls in range(num_classes):
        true_positive_pixels[cls] = np.sum((pred == cls) * (target == cls))
        positive_pixels[cls] += np.sum(pred == cls)
        negative_pixels[cls] += np.sum(target != cls)

    # 计算Precision、PA、MIoU、FWIoU、F1
    precision = np.zeros(num_classes)
    for cls in range(num_classes):
        if positive_pixels[cls] > 0:
            precision[cls] = true_positive_pixels[cls] / positive_pixels[cls]

    pa = np.sum(true_positive_pixels) / total_pixels

    iou = np.zeros(num_classes)
    for cls in range(num_classes):
        union = positive_pixels[cls] + negative_pixels[cls] - true_positive_pixels[cls]
        if union > 0:
            iou[cls] = true_positive_pixels[cls] / union

    miou = np.mean(iou)

    freq = positive_pixels / np.sum(positive_pixels)
    fwiou = np.sum(freq * iou)

    f1 = np.zeros(num_classes)
    for cls in range(num_classes):
        precision_cls = precision[cls]
        recall_cls = true_positive_pixels[cls] / (positive_pixels[cls] + negative_pixels[cls] - true_positive_pixels[cls])
        if precision_cls + recall_cls > 0:
            f1[cls] = 2 * precision_cls * recall_cls / (precision_cls + recall_cls)

    return np.mean(precision), pa, miou, fwiou, np.mean(f1)



def save_metrics_to_excel(metrics, file_name='result/metrics.xlsx', sheet_name='Sheet'):
    """
    将评估指标保存到Excel文件中。

    参数:
    metrics (list): 一个列表,每个元素是一个元组,包含批次号和对应的评估指标值。
    file_name (str, optional): Excel文件名,默认为'metrics.xlsx'。
    sheet_name (str, optional): 工作表名称,默认为'Sheet'。
    """
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()
    # 获取活动工作表
    worksheet = workbook.active
    worksheet.title = sheet_name

    # 写入表头
    worksheet['A1'] = 'Picture number'
    worksheet['B1'] = 'Precision'
    worksheet['C1'] = 'Pixel Accuracy'
    worksheet['D1'] = 'Mean IoU'
    worksheet['E1'] = 'Frequency Weighted IoU'
    worksheet['F1'] = 'F1 Score'

    # 写入评估指标数据
    for row, (batch, precision, pa, miou, fwiou, f1) in enumerate(metrics, start=2):
        worksheet.cell(row=row, column=1, value=batch)
        worksheet.cell(row=row, column=2, value=precision)
        worksheet.cell(row=row, column=3, value=pa)
        worksheet.cell(row=row, column=4, value=miou)
        worksheet.cell(row=row, column=5, value=fwiou)
        worksheet.cell(row=row, column=6, value=f1)

    # 保存Excel文件
    workbook.save(file_name)

if __name__ == '__main__':
    ## 设置参数--------------------------------------------------------------------------------------------------------------------------------
    num_classes = 10 + 1
    _input = 'data\测试集'
    encoder = load_encoder("params/encoder.pkl")
    weights = 'params/60.pth'
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    save_path = 'result'
    net = UNet(num_classes).to(device)
    if os.path.exists(weights):
        net.load_state_dict(torch.load(weights))
        print('successfully')
    else:
        print('no loading')
    data_loader = DataLoader(MyDataset(_input, encoder), batch_size=1, shuffle=False)
    ## 预测----------------------------------------------------------------------------------------------------------------------------------
    metrics = []
    # 禁用梯度计算以加速推理
    # 初始化列表存储预测结果
    all_preds = []
    all_targets = []
    # 遍历DataLoader
    for i, (image, segment_image) in enumerate(tqdm.tqdm(data_loader)):
        # 将输入和目标数据移动到相同设备(CPU或GPU)
        image, segment_image = image.to(device), segment_image.to(device)
        # 前向传播获取预测结果
        outputs = net(image)
        # 后处理可视化
        _image = image[0].cpu().numpy()
        _segment_image = encoder.decode(segment_image[0])
        _out_image = encoder.decode(outputs[0])
        plt.figure(figsize=(10, 10))
        plt.imshow(_out_image)
        # 垂直方向拼接
        _image = np.swapaxes(_image, 0, 2) * 255
        img_vstack = np.vstack((_segment_image, _image, _out_image))
        # img_vstack = _out_image
        # 保存垂直拼接图像
        img_vstack = cv2.cvtColor(img_vstack, cv2.COLOR_BGR2RGB)
        cv2.imwrite(f'{save_path}/{i}.png', img_vstack)
        precision, pa, miou, fwiou, f1 = calculate_sem_seg_metrics(outputs[0], segment_image[0], num_classes)
        # 将评估指标添加到列表中
        metrics.append((i, precision, pa, miou, fwiou, f1))

    # 保存评估指标到Excel文件
    save_metrics_to_excel(metrics)